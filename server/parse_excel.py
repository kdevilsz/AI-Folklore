import openpyxl
import json
import os

EXCEL_PATH = r"c:\Users\Bhaswati Sikdar\Documents\data\Assamese_Folklore_Dataset (1).xlsx"
OUT_PATH = r"c:\Users\Bhaswati Sikdar\Documents\data\js\data.js"

def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Assuming Proverbs is the active or first sheet, or named "Proverbs"
    proverbs = []
    if "Proverbs" in wb.sheetnames:
        sheet = wb["Proverbs"]
    else:
        sheet = wb.active
        
    # Read headers
    headers = [cell.value for cell in sheet[1]]
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue # skip empty
        
        proverb = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                key = str(h).strip().lower().replace(" ", "_")
                val = row[i] if row[i] is not None else ""
                proverb[key] = str(val).strip()
        
        # Format for our frontend
        formatted = {
            "id": f"p_{len(proverbs)+1}",
            "assamese": proverb.get('assamese', proverb.get('assamese_text', proverb.get('text', ''))),
            "transliteration": proverb.get('transliteration', ''),
            "meaning": proverb.get('meaning', proverb.get('english_translation', '')),
            "themes": proverb.get('themes', '').split(',') if proverb.get('themes') else []
        }
        proverbs.append(formatted)

    js_content = f"export const proverbs = {json.dumps(proverbs, indent=4, ensure_ascii=False)};\n"
    
    # Keep the dummy folktales for now, since RAG handles the heavy lifting
    js_content += """
export const folktales = [
    { id: "f1", title: "Tejimola", summary: "A tragic tale of a stepdaughter who turns into various plants. (Check Chat for full details!)" },
    { id: "f2", title: "Burhi Aair Xadhu", summary: "Grandma's tales compilation." }
];
"""

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        f.write(js_content)
        
    print(f"Extracted {len(proverbs)} proverbs and wrote to {OUT_PATH}")

if __name__ == "__main__":
    parse_excel()
